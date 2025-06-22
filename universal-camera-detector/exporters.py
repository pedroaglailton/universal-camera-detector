# universal_camera_detector/exporters.py

import io
import pandas as pd
import openpyxl
from openpyxl.drawing.image import Image as ExcelImage
import os
import tempfile
import base64
from typing import List, Dict

def export_to_excel_with_images(cameras: List[Dict], filename: str) -> bytes:
    """Exporta dados para Excel com imagens incorporadas"""
    try:
        import openpyxl
    except ImportError:
        raise ImportError("Instale o pacote openpyxl: pip install openpyxl")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Câmeras Descobertas"

    # Cabeçalhos
    headers = ['IP', 'Marca', 'Modelo', 'Serial', 'IP Atual', 'Máscara', 'Gateway', 'DHCP', 'Status', 'Thumbnail']
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)

    temp_files = []
    row = 2

    for camera in cameras:
        ws.cell(row=row, column=1, value=camera['IP'])
        ws.cell(row=row, column=2, value=camera['Marca'])
        ws.cell(row=row, column=3, value=camera['Modelo'])
        ws.cell(row=row, column=4, value=camera['Serial'])
        ws.cell(row=row, column=5, value=camera['IP Atual'])
        ws.cell(row=row, column=6, value=camera['Máscara'])
        ws.cell(row=row, column=7, value=camera['Gateway'])
        ws.cell(row=row, column=8, value=camera['DHCP'])
        ws.cell(row=row, column=9, value=camera['Status'])

        if camera.get('Thumbnail_Bytes'):
            try:
                tmp_file = tempfile.NamedTemporaryFile(suffix='.jpg', delete=False)
                tmp_file.write(camera['Thumbnail_Bytes'])
                tmp_file.close()
                img = ExcelImage(tmp_file.name)
                img.width = 100
                img.height = 75
                ws.add_image(img, f"J{row}")
                ws.row_dimensions[row].height = 60
                temp_files.append(tmp_file.name)
            except Exception as e:
                logger.error(f"Erro ao adicionar imagem para {camera['IP']}: {e}")

        row += 1

    # Ajusta tamanho das colunas
    for col in range(1, 10):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15

    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)

    # Limpa temporários
    for tmp_path in temp_files:
        try:
            os.unlink(tmp_path)
        except:
            pass

    return excel_buffer.getvalue()


def export_to_csv_with_base64(cameras: List[Dict]) -> bytes:
    """Exporta dados para CSV com imagens em base64"""
    export_data = []
    for camera in cameras:
        cam_copy = camera.copy()
        cam_copy.pop('Thumbnail_Bytes', None)
        cam_copy.pop('camera_info', None)
        export_data.append(cam_copy)
    
    df = pd.DataFrame(export_data)
    return df.to_csv(index=False).encode('utf-8')
